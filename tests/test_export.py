import io

from openpyxl import load_workbook

from app.models import Van, Driver


class TestExport:
    def _seed_and_assign(self, client, db_session, headers):
        v = Van(code="VAN-001", description="Test")
        d = Driver(employee_id="EMP-001", name="John")
        db_session.add_all([v, d])
        db_session.commit()

        client.post("/api/assignments", json={
            "assignment_date": "2026-01-05",
            "van_id": v.id,
            "driver_id": d.id,
        }, headers=headers)
        return v, d

    def test_export_daily_xlsx(self, client, db_session, operator_headers):
        self._seed_and_assign(client, db_session, operator_headers)

        resp = client.get("/api/export/daily?target_date=2026-01-05", headers=operator_headers)
        assert resp.status_code == 200
        assert "spreadsheetml" in resp.headers["content-type"]

        wb = load_workbook(io.BytesIO(resp.content))
        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        assert len(rows) == 1
        assert rows[0][0] == "2026-01-05"
        assert rows[0][2] == "EMP-001"
        assert rows[0][4] == "VAN-001"
        assert rows[0][7] == "Assigned"

    def test_export_weekly_xlsx(self, client, db_session, operator_headers):
        self._seed_and_assign(client, db_session, operator_headers)

        resp = client.get("/api/export/weekly?week=2", headers=operator_headers)
        assert resp.status_code == 200

        wb = load_workbook(io.BytesIO(resp.content))
        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        assert len(rows) == 1
        assert rows[0][1] == 2

    def test_export_empty_day(self, client, db_session, operator_headers):
        resp = client.get("/api/export/daily?target_date=2026-03-15", headers=operator_headers)
        assert resp.status_code == 200

        wb = load_workbook(io.BytesIO(resp.content))
        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        assert len(rows) == 0

    def test_export_xlsx_has_headers(self, client, db_session, operator_headers):
        resp = client.get("/api/export/daily?target_date=2026-01-05", headers=operator_headers)
        wb = load_workbook(io.BytesIO(resp.content))
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        assert "Date" in headers
        assert "Week #" in headers
        assert "Driver ID" in headers
        assert "Van (Plate)" in headers
        assert "Status" in headers

    def test_export_requires_auth(self, client):
        resp = client.get("/api/export/daily?target_date=2026-01-05")
        assert resp.status_code == 401
