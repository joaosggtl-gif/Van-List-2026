import io

from openpyxl import load_workbook

from app.models import Van, Driver


class TestPartialAssignments:
    """Test driver-only, van-only, pair, unpair, and bulk workflows."""

    def _seed(self, db):
        v1 = Van(code="VAN-001", description="Test Van 1")
        v2 = Van(code="VAN-002", description="Test Van 2")
        d1 = Driver(employee_id="EMP-001", name="Driver One")
        d2 = Driver(employee_id="EMP-002", name="Driver Two")
        db.add_all([v1, v2, d1, d2])
        db.commit()
        return v1, v2, d1, d2

    def test_create_driver_only(self, client, db_session, operator_headers):
        v1, v2, d1, d2 = self._seed(db_session)
        resp = client.post("/api/assignments", json={
            "assignment_date": "2026-01-05",
            "driver_id": d1.id,
        }, headers=operator_headers)
        assert resp.status_code == 201
        data = resp.json()
        assert data["driver"]["employee_id"] == "EMP-001"
        assert data["van"] is None
        assert data["van_id"] is None

    def test_create_van_only(self, client, db_session, operator_headers):
        v1, v2, d1, d2 = self._seed(db_session)
        resp = client.post("/api/assignments", json={
            "assignment_date": "2026-01-05",
            "van_id": v1.id,
        }, headers=operator_headers)
        assert resp.status_code == 201
        data = resp.json()
        assert data["van"]["code"] == "VAN-001"
        assert data["driver"] is None
        assert data["driver_id"] is None

    def test_create_neither_fails_422(self, client, db_session, operator_headers):
        self._seed(db_session)
        resp = client.post("/api/assignments", json={
            "assignment_date": "2026-01-05",
        }, headers=operator_headers)
        assert resp.status_code == 422

    def test_create_full_still_works(self, client, db_session, operator_headers):
        v1, v2, d1, d2 = self._seed(db_session)
        resp = client.post("/api/assignments", json={
            "assignment_date": "2026-01-05",
            "van_id": v1.id,
            "driver_id": d1.id,
        }, headers=operator_headers)
        assert resp.status_code == 201
        data = resp.json()
        assert data["van"]["code"] == "VAN-001"
        assert data["driver"]["employee_id"] == "EMP-001"

    def test_pair_assignments(self, client, db_session, operator_headers):
        v1, v2, d1, d2 = self._seed(db_session)

        # Create driver-only
        resp_d = client.post("/api/assignments", json={
            "assignment_date": "2026-01-05",
            "driver_id": d1.id,
        }, headers=operator_headers)
        assert resp_d.status_code == 201
        driver_asgn_id = resp_d.json()["id"]

        # Create van-only
        resp_v = client.post("/api/assignments", json={
            "assignment_date": "2026-01-05",
            "van_id": v1.id,
        }, headers=operator_headers)
        assert resp_v.status_code == 201
        van_asgn_id = resp_v.json()["id"]

        # Pair them
        resp_pair = client.post("/api/assignments/pair", json={
            "driver_assignment_id": driver_asgn_id,
            "van_assignment_id": van_asgn_id,
        }, headers=operator_headers)
        assert resp_pair.status_code == 200
        data = resp_pair.json()
        assert data["van"]["code"] == "VAN-001"
        assert data["driver"]["employee_id"] == "EMP-001"
        assert data["van_id"] is not None
        assert data["driver_id"] is not None

        # Van-only row should be deleted (only 1 assignment left)
        resp_list = client.get("/api/assignments?date_from=2026-01-05", headers=operator_headers)
        assert len(resp_list.json()) == 1

    def test_unpair_assignment(self, client, db_session, operator_headers):
        v1, v2, d1, d2 = self._seed(db_session)

        # Create full assignment
        resp = client.post("/api/assignments", json={
            "assignment_date": "2026-01-05",
            "van_id": v1.id,
            "driver_id": d1.id,
        }, headers=operator_headers)
        assert resp.status_code == 201
        asgn_id = resp.json()["id"]

        # Unpair
        resp_unpair = client.post(f"/api/assignments/{asgn_id}/unpair", headers=operator_headers)
        assert resp_unpair.status_code == 200
        data = resp_unpair.json()
        assert "driver_assignment_id" in data
        assert "van_assignment_id" in data

        # Should now have 2 assignments
        resp_list = client.get("/api/assignments?date_from=2026-01-05", headers=operator_headers)
        assignments = resp_list.json()
        assert len(assignments) == 2
        # One should be driver-only, one van-only
        has_driver_only = any(a["driver_id"] and not a["van_id"] for a in assignments)
        has_van_only = any(a["van_id"] and not a["driver_id"] for a in assignments)
        assert has_driver_only
        assert has_van_only

    def test_unpair_already_partial_fails(self, client, db_session, operator_headers):
        v1, v2, d1, d2 = self._seed(db_session)

        # Create driver-only
        resp = client.post("/api/assignments", json={
            "assignment_date": "2026-01-05",
            "driver_id": d1.id,
        }, headers=operator_headers)
        assert resp.status_code == 201
        asgn_id = resp.json()["id"]

        # Unpair should fail
        resp_unpair = client.post(f"/api/assignments/{asgn_id}/unpair", headers=operator_headers)
        assert resp_unpair.status_code == 400

    def test_pair_wrong_types_fails(self, client, db_session, operator_headers):
        v1, v2, d1, d2 = self._seed(db_session)

        # Create two driver-only assignments
        resp_d1 = client.post("/api/assignments", json={
            "assignment_date": "2026-01-05",
            "driver_id": d1.id,
        }, headers=operator_headers)
        resp_d2 = client.post("/api/assignments", json={
            "assignment_date": "2026-01-05",
            "driver_id": d2.id,
        }, headers=operator_headers)

        # Try to pair two driver-only assignments (second has no van)
        resp_pair = client.post("/api/assignments/pair", json={
            "driver_assignment_id": resp_d1.json()["id"],
            "van_assignment_id": resp_d2.json()["id"],
        }, headers=operator_headers)
        assert resp_pair.status_code == 400

    def test_pair_different_dates_fails(self, client, db_session, operator_headers):
        v1, v2, d1, d2 = self._seed(db_session)

        resp_d = client.post("/api/assignments", json={
            "assignment_date": "2026-01-05",
            "driver_id": d1.id,
        }, headers=operator_headers)
        resp_v = client.post("/api/assignments", json={
            "assignment_date": "2026-01-06",
            "van_id": v1.id,
        }, headers=operator_headers)

        resp_pair = client.post("/api/assignments/pair", json={
            "driver_assignment_id": resp_d.json()["id"],
            "van_assignment_id": resp_v.json()["id"],
        }, headers=operator_headers)
        assert resp_pair.status_code == 400
        assert "same date" in resp_pair.json()["detail"]

    def test_available_vans_with_partial_entries(self, client, db_session, operator_headers):
        """Van-only assignment should exclude van from available list."""
        v1, v2, d1, d2 = self._seed(db_session)

        # Create van-only assignment for v1
        client.post("/api/assignments", json={
            "assignment_date": "2026-01-05",
            "van_id": v1.id,
        }, headers=operator_headers)

        resp = client.get("/api/assignments/available-vans?assignment_date=2026-01-05", headers=operator_headers)
        assert resp.status_code == 200
        codes = [v["code"] for v in resp.json()]
        assert "VAN-001" not in codes
        assert "VAN-002" in codes

    def test_available_drivers_with_partial_entries(self, client, db_session, operator_headers):
        """Driver-only assignment should exclude driver from available list."""
        v1, v2, d1, d2 = self._seed(db_session)

        # Create driver-only assignment for d1
        client.post("/api/assignments", json={
            "assignment_date": "2026-01-05",
            "driver_id": d1.id,
        }, headers=operator_headers)

        resp = client.get("/api/assignments/available-drivers?assignment_date=2026-01-05", headers=operator_headers)
        assert resp.status_code == 200
        ids = [d["employee_id"] for d in resp.json()]
        assert "EMP-001" not in ids
        assert "EMP-002" in ids

    def test_driver_only_duplicate_same_day_blocked(self, client, db_session, operator_headers):
        v1, v2, d1, d2 = self._seed(db_session)

        resp1 = client.post("/api/assignments", json={
            "assignment_date": "2026-01-05",
            "driver_id": d1.id,
        }, headers=operator_headers)
        assert resp1.status_code == 201

        resp2 = client.post("/api/assignments", json={
            "assignment_date": "2026-01-05",
            "driver_id": d1.id,
        }, headers=operator_headers)
        assert resp2.status_code == 409

    def test_van_only_duplicate_same_day_blocked(self, client, db_session, operator_headers):
        v1, v2, d1, d2 = self._seed(db_session)

        resp1 = client.post("/api/assignments", json={
            "assignment_date": "2026-01-05",
            "van_id": v1.id,
        }, headers=operator_headers)
        assert resp1.status_code == 201

        resp2 = client.post("/api/assignments", json={
            "assignment_date": "2026-01-05",
            "van_id": v1.id,
        }, headers=operator_headers)
        assert resp2.status_code == 409


class TestExportMixedStates:
    """Test export with driver-only, van-only, and paired assignments."""

    def test_export_mixed_states(self, client, db_session, operator_headers):
        v = Van(code="VAN-001", description="Test")
        d = Driver(employee_id="EMP-001", name="John")
        v2 = Van(code="VAN-002", description="Test 2")
        d2 = Driver(employee_id="EMP-002", name="Jane")
        db_session.add_all([v, d, v2, d2])
        db_session.commit()

        # Full assignment
        client.post("/api/assignments", json={
            "assignment_date": "2026-01-05",
            "van_id": v.id,
            "driver_id": d.id,
        }, headers=operator_headers)

        # Driver-only
        client.post("/api/assignments", json={
            "assignment_date": "2026-01-05",
            "driver_id": d2.id,
        }, headers=operator_headers)

        # Van-only
        client.post("/api/assignments", json={
            "assignment_date": "2026-01-05",
            "van_id": v2.id,
        }, headers=operator_headers)

        resp = client.get("/api/export/daily?target_date=2026-01-05", headers=operator_headers)
        assert resp.status_code == 200

        wb = load_workbook(io.BytesIO(resp.content))
        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        assert len(rows) == 3

        statuses = [r[7] for r in rows]
        assert "Assigned" in statuses
        assert "Driver Only" in statuses
        assert "Van Only" in statuses

        # Check driver-only row has empty van fields
        driver_only_row = [r for r in rows if r[7] == "Driver Only"][0]
        assert driver_only_row[2] == "EMP-002"  # driver id
        assert driver_only_row[4] in ("", None)  # van plate empty

        # Check van-only row has empty driver fields
        van_only_row = [r for r in rows if r[7] == "Van Only"][0]
        assert van_only_row[4] == "VAN-002"  # van plate
        assert van_only_row[2] in ("", None)  # driver id empty


class TestBulkUploadAssignments:
    """Test bulk upload of drivers and vans for a specific day."""

    def test_bulk_upload_drivers_csv(self, client, db_session, operator_headers):
        csv_content = b"employee_id,name\nEMP-001,Driver One\nEMP-002,Driver Two\n"
        resp = client.post(
            "/api/assignments/bulk-upload-drivers?assignment_date=2026-01-05",
            files={"file": ("drivers.csv", csv_content, "text/csv")},
            headers=operator_headers,
        )
        assert resp.status_code == 200
        data = resp.json()
        assert data["assignments_created"] == 2

        # Verify assignments were created
        resp_list = client.get("/api/assignments?date_from=2026-01-05", headers=operator_headers)
        assignments = resp_list.json()
        assert len(assignments) == 2
        # All should be driver-only
        for a in assignments:
            assert a["driver_id"] is not None
            assert a["van_id"] is None

    def test_bulk_upload_vans_csv(self, client, db_session, operator_headers):
        csv_content = b"code,description\nVAN-001,Test Van 1\nVAN-002,Test Van 2\n"
        resp = client.post(
            "/api/assignments/bulk-upload-vans?assignment_date=2026-01-05",
            files={"file": ("vans.csv", csv_content, "text/csv")},
            headers=operator_headers,
        )
        assert resp.status_code == 200
        data = resp.json()
        assert data["assignments_created"] == 2

        # Verify assignments were created
        resp_list = client.get("/api/assignments?date_from=2026-01-05", headers=operator_headers)
        assignments = resp_list.json()
        assert len(assignments) == 2
        # All should be van-only
        for a in assignments:
            assert a["van_id"] is not None
            assert a["driver_id"] is None

    def test_bulk_upload_skips_already_assigned(self, client, db_session, operator_headers):
        # Pre-create a van and assign it
        v = Van(code="VAN-001", description="Test")
        db_session.add(v)
        db_session.commit()

        client.post("/api/assignments", json={
            "assignment_date": "2026-01-05",
            "van_id": v.id,
        }, headers=operator_headers)

        # Upload CSV that includes VAN-001 again
        csv_content = b"code,description\nVAN-001,Test Van 1\nVAN-002,Test Van 2\n"
        resp = client.post(
            "/api/assignments/bulk-upload-vans?assignment_date=2026-01-05",
            files={"file": ("vans.csv", csv_content, "text/csv")},
            headers=operator_headers,
        )
        assert resp.status_code == 200
        data = resp.json()
        assert data["assignments_created"] == 1  # Only VAN-002
        assert data["assignments_skipped"] == 1  # VAN-001 was already assigned
