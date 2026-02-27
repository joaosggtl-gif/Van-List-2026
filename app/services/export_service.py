import io
from datetime import date, timedelta

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from sqlalchemy.orm import Session, joinedload

from app.models import DailyAssignment, Van
from app.routes.pages import short_name
from app.services.week_service import get_week_number, get_week_days


HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def _style_header(ws, row=1):
    for cell in ws[row]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER


def _auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 3, 40)


def _assignment_row(a: DailyAssignment) -> list:
    # Determine status
    if a.van_id is not None and a.driver_id is not None:
        status = "Assigned"
    elif a.driver_id is not None:
        status = "Driver Only"
    else:
        status = "Van Only"

    return [
        a.assignment_date.isoformat(),
        get_week_number(a.assignment_date),
        a.driver.employee_id if a.driver else "",
        short_name(a.driver.name) if a.driver else "",
        a.van.code if a.van else "",
        (a.van.description or "") if a.van else "",
        (a.van.operational_status or "") if a.van else "",
        status,
        a.created_at.strftime("%Y-%m-%d %H:%M:%S"),
        a.updated_at.strftime("%Y-%m-%d %H:%M:%S"),
        a.notes or "",
    ]


HEADERS = [
    "Date", "Week #", "Driver ID", "Driver Name",
    "Van (Plate)", "Van Description", "Operational Status", "Status",
    "Created At", "Updated At", "Notes",
]


def export_daily_xlsx(db: Session, target_date: date) -> bytes:
    """Export a single day's assignments to XLSX."""
    assignments = (
        db.query(DailyAssignment)
        .options(joinedload(DailyAssignment.van), joinedload(DailyAssignment.driver))
        .filter(DailyAssignment.assignment_date == target_date)
        .order_by(DailyAssignment.id)
        .all()
    )

    wb = Workbook()
    ws = wb.active
    ws.title = f"Assignments {target_date.isoformat()}"

    ws.append(HEADERS)
    _style_header(ws)

    for a in assignments:
        ws.append(_assignment_row(a))

    _auto_width(ws)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


FREE_FILL = PatternFill(start_color="D1E7DD", end_color="D1E7DD", fill_type="solid")
FREE_FONT = Font(color="0F5132")
VOR_FILL = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")


def export_daily_simple_xlsx(db: Session, target_date: date) -> bytes:
    """Export a single day's van list as a simple Van Reg / Driver Name XLSX.

    Order: assigned vans first, then Free vans, then VOR vans at the bottom.
    """
    # All active vans
    all_vans = db.query(Van).filter(Van.active == True).order_by(Van.code).all()

    # Assignments for the target date (van-based)
    assignments = (
        db.query(DailyAssignment)
        .options(joinedload(DailyAssignment.van), joinedload(DailyAssignment.driver))
        .filter(
            DailyAssignment.assignment_date == target_date,
            DailyAssignment.van_id.isnot(None),
        )
        .all()
    )

    # Index assignments by van_id
    assign_by_van = {}
    for a in assignments:
        assign_by_van[a.van_id] = a

    # Categorise vans
    assigned_rows = []
    free_rows = []
    vor_rows = []

    for van in all_vans:
        a = assign_by_van.get(van.id)
        if a and a.driver_id and a.driver:
            assigned_rows.append((van.code, short_name(a.driver.name), "assigned"))
        elif a and not a.driver_id:
            vor_rows.append((van.code, "VOR", "vor"))
        elif van.operational_status == "GROUNDED":
            vor_rows.append((van.code, "VOR", "vor"))
        else:
            free_rows.append((van.code, "Free", "free"))

    wb = Workbook()
    ws = wb.active
    ws.title = target_date.strftime("%A %d-%m")

    ws.append(["Van Reg", "Driver Name"])
    _style_header(ws)

    for code, driver, status in assigned_rows + free_rows + vor_rows:
        ws.append([code, driver])
        row_num = ws.max_row
        if status == "free":
            for col in (1, 2):
                cell = ws.cell(row=row_num, column=col)
                cell.fill = FREE_FILL
                cell.font = FREE_FONT
        elif status == "vor":
            for col in (1, 2):
                cell = ws.cell(row=row_num, column=col)
                cell.fill = VOR_FILL
                cell.font = VOR_FONT

    _auto_width(ws)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_weekly_xlsx(db: Session, week_number: int) -> bytes:
    """Export a full week's assignments to XLSX."""
    days = get_week_days(week_number)
    start_date = days[0]
    end_date = days[-1]

    assignments = (
        db.query(DailyAssignment)
        .options(joinedload(DailyAssignment.van), joinedload(DailyAssignment.driver))
        .filter(
            DailyAssignment.assignment_date >= start_date,
            DailyAssignment.assignment_date <= end_date,
        )
        .order_by(DailyAssignment.assignment_date, DailyAssignment.id)
        .all()
    )

    wb = Workbook()
    ws = wb.active
    ws.title = f"Week {week_number}"

    ws.append(HEADERS)
    _style_header(ws)

    for a in assignments:
        ws.append(_assignment_row(a))

    _auto_width(ws)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


DAY_NAMES = ["Sunday", "Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday"]

VOR_FONT = Font(bold=True, color="DC3545")


def export_period_xlsx(db: Session, start_date: date, end_date: date) -> bytes:
    """Export a date range as XLSX with one tab per week, matching the weekly grid layout."""
    # Determine which weeks are spanned
    start_week = get_week_number(start_date)
    end_week = get_week_number(end_date)

    # Load all active vans
    all_vans = db.query(Van).filter(Van.active == True).all()

    # Load all assignments in the date range
    assignments = (
        db.query(DailyAssignment)
        .options(joinedload(DailyAssignment.van), joinedload(DailyAssignment.driver))
        .filter(
            DailyAssignment.assignment_date >= start_date,
            DailyAssignment.assignment_date <= end_date,
        )
        .order_by(DailyAssignment.assignment_date, DailyAssignment.id)
        .all()
    )

    # Index assignments by (van_id, date_iso)
    assign_lookup = {}
    vans_with_drivers = set()
    for a in assignments:
        if a.van_id:
            assign_lookup[(a.van_id, a.assignment_date.isoformat())] = a
            if a.driver_id:
                vans_with_drivers.add(a.van_id)

    # Sort vans: OPERATIONAL first by code, then GROUNDED with drivers, then GROUNDED without
    all_vans.sort(key=lambda v: (
        v.operational_status == 'GROUNDED',
        not (v.id in vans_with_drivers) if v.operational_status == 'GROUNDED' else False,
        v.code,
    ))

    wb = Workbook()
    # Remove the default sheet — we'll create our own
    wb.remove(wb.active)

    for week_num in range(start_week, end_week + 1):
        days = get_week_days(week_num)
        week_start = days[0]
        week_end = days[-1]

        tab_name = f"Week {week_num} ({week_start.strftime('%d-%m')}\u2013{week_end.strftime('%d-%m')})"
        # Worksheet names max 31 chars
        if len(tab_name) > 31:
            tab_name = tab_name[:31]

        ws = wb.create_sheet(title=tab_name)

        # Header row: Van | Sunday (dd/mm) | Monday (dd/mm) | ... | Saturday (dd/mm)
        headers = ["Van"]
        for i, day in enumerate(days):
            headers.append(f"{DAY_NAMES[i]}\n{day.strftime('%d/%m')}")
        ws.append(headers)
        _style_header(ws)
        # Wrap text in header for the date line
        for cell in ws[1]:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # One row per van
        for van in all_vans:
            row_data = [van.code]
            row_cells_vor = []  # track which columns need VOR styling
            for i, day in enumerate(days):
                day_iso = day.isoformat()
                a = assign_lookup.get((van.id, day_iso))
                if a and a.driver_id and a.driver:
                    row_data.append(short_name(a.driver.name))
                    row_cells_vor.append(False)
                elif a and not a.driver_id:
                    # Van-only assignment = VOR
                    row_data.append("VOR")
                    row_cells_vor.append(True)
                elif van.operational_status == 'GROUNDED':
                    # Grounded van with no assignment = VOR
                    row_data.append("VOR")
                    row_cells_vor.append(True)
                else:
                    row_data.append("")
                    row_cells_vor.append(False)
            ws.append(row_data)

            # Apply VOR styling (red bold)
            current_row = ws.max_row
            for col_idx, is_vor in enumerate(row_cells_vor):
                if is_vor:
                    ws.cell(row=current_row, column=col_idx + 2).font = VOR_FONT

        _auto_width(ws)
        # Set header row height for wrapped text
        ws.row_dimensions[1].height = 35

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
