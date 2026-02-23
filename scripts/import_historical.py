"""Import historical van assignments from Van Use Tracker DRR1.xlsx.

Reads each week sheet, extracts van registrations and driver assignments,
and inserts them into the historical_assignments table.

Usage:
    python -m scripts.import_historical
"""
import sys
import os
from datetime import date, timedelta, datetime

import openpyxl

# Add parent to path so we can import app
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from app.database import SessionLocal
from app.models import HistoricalAssignment

XLSX_PATH = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
    "Van Use Tracker DRR1.xlsx",
)

# Column letters for each day of the week (Sun=D, Mon=G, ... Sat=V)
DAY_COLUMNS = ["D", "G", "J", "M", "P", "S", "V"]
DAY_OFFSETS = [0, 1, 2, 3, 4, 5, 6]  # offset from Sunday

# Sheets to skip
SKIP_SHEETS = {"Names"}


def col_letter_to_index(letter: str) -> int:
    """Convert column letter to 1-based index."""
    result = 0
    for c in letter.upper():
        result = result * 26 + (ord(c) - ord("A") + 1)
    return result


def parse_sunday_date(ws) -> date | None:
    """Extract the Sunday date from row 2, column D."""
    cell_val = ws.cell(row=2, column=col_letter_to_index("D")).value
    if cell_val is None:
        return None
    if isinstance(cell_val, datetime):
        return cell_val.date()
    if isinstance(cell_val, date):
        return cell_val
    # Try parsing string formats like '15/2/2026'
    s = str(cell_val).strip()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%m/%d/%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def is_free_or_empty(val: str | None) -> bool:
    """Check if cell value means no assignment."""
    if val is None:
        return True
    s = str(val).strip().lower()
    return s in ("", "free", "n/a", "-", "—")


def is_vor(val: str | None) -> bool:
    """Check if cell value indicates VOR (Van Out of Road)."""
    if val is None:
        return False
    return str(val).strip().upper() == "VOR"


def import_xlsx():
    print(f"Loading workbook: {XLSX_PATH}")
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True, read_only=True)

    db = SessionLocal()
    total_inserted = 0
    total_skipped = 0
    total_vor = 0

    try:
        for sheet_name in wb.sheetnames:
            if sheet_name.strip() in SKIP_SHEETS:
                continue

            ws = wb[sheet_name]
            sunday = parse_sunday_date(ws)
            if sunday is None:
                print(f"  SKIP {sheet_name}: could not parse Sunday date")
                continue

            print(f"\n=== {sheet_name} (Sunday: {sunday}) ===")

            sheet_inserted = 0
            sheet_skipped = 0
            seen = set()  # track (date, van_reg) to avoid duplicates within batch

            for row in ws.iter_rows(min_row=3, values_only=False):
                # Column C = registration
                van_reg_cell = row[col_letter_to_index("C") - 1]
                van_reg = van_reg_cell.value
                if not van_reg or not str(van_reg).strip():
                    continue
                van_reg = str(van_reg).strip().upper()

                for day_col, day_offset in zip(DAY_COLUMNS, DAY_OFFSETS):
                    col_idx = col_letter_to_index(day_col) - 1
                    if col_idx >= len(row):
                        continue
                    cell_val = row[col_idx].value
                    assignment_date = sunday + timedelta(days=day_offset)

                    driver_name = None
                    vor = False

                    if is_vor(cell_val):
                        vor = True
                    elif is_free_or_empty(cell_val):
                        # No assignment - skip (don't store empty cells)
                        continue
                    else:
                        driver_name = str(cell_val).strip()

                    key = (assignment_date, van_reg)
                    if key in seen:
                        sheet_skipped += 1
                        continue
                    seen.add(key)

                    record = HistoricalAssignment(
                        assignment_date=assignment_date,
                        van_reg=van_reg,
                        driver_name=driver_name,
                        is_vor=vor,
                    )
                    db.add(record)
                    sheet_inserted += 1
                    if vor:
                        total_vor += 1

            db.commit()
            total_inserted += sheet_inserted
            total_skipped += sheet_skipped
            print(f"  Inserted: {sheet_inserted}, Skipped (duplicate): {sheet_skipped}")

        print(f"\n{'='*50}")
        print(f"TOTAL inserted: {total_inserted}")
        print(f"TOTAL VOR entries: {total_vor}")
        print(f"TOTAL skipped (duplicate): {total_skipped}")

    except Exception as e:
        db.rollback()
        print(f"ERROR: {e}")
        raise
    finally:
        db.close()
        wb.close()


if __name__ == "__main__":
    import_xlsx()
